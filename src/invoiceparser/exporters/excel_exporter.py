"""Экспорт в Excel"""
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from ..core.config import Config
from ..core.models import InvoiceData

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Экспортер в Excel формат"""

    def __init__(self, config: Config):
        """
        Инициализация экспортера

        Args:
            config: Конфигурация приложения
        """
        self.config = config
        self.output_dir = config.output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, document_path: Path, invoice_data: InvoiceData, raw_data: Optional[Dict[str, Any]] = None) -> Path:
        """
        Экспорт данных счета в Excel

        Args:
            document_path: Путь к исходному документу
            invoice_data: Данные счета
            raw_data: Исходные данные в формате dict (для структурированного экспорта)

        Returns:
            Путь к созданному Excel файлу
        """
        # Генерируем имя файла на основе исходного документа
        filename_base = document_path.stem
        output_path = self.output_dir / f"{filename_base}.xlsx"

        wb = Workbook()
        # Удаляем стандартный лист если он существует
        default_sheet_name = self.config.excel_default_sheet_name
        if default_sheet_name in wb.sheetnames:
            wb.remove(wb[default_sheet_name])

        # Если есть raw_data, используем структурированный экспорт
        if raw_data:
            self._export_structured(wb, raw_data)
        else:
            # Fallback на старый формат
            self._export_simple(wb, invoice_data)

        wb.save(output_path)
        logger.info(f"Excel exported: {output_path.name}")
        return output_path

    def _export_structured(self, wb: Workbook, data: Dict[str, Any]):
        """
        Структурированный экспорт данных в понятном формате

        Args:
            wb: Workbook
            data: Данные в формате dict
        """
        # Header sheet
        header_sheet_name = self.config.excel_sheet_header_name
        ws1 = wb.create_sheet(header_sheet_name)

        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        value_alignment = Alignment(wrap_text=True, vertical="top")

        row = 1

        # 1. Інформація про документ
        ws1.merge_cells(f'A{row}:B{row}')
        ws1[f'A{row}'] = "Інформація про документ"
        ws1[f'A{row}'].fill = section_fill
        ws1[f'A{row}'].font = section_font
        row += 1

        # Заголовки колонок
        ws1[f'A{row}'] = self.config.excel_header_field_column
        ws1[f'B{row}'] = self.config.excel_header_value_column
        ws1[f'A{row}'].fill = header_fill
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].fill = header_fill
        ws1[f'B{row}'].font = header_font
        row += 1

        # Данные документа
        doc_info = data.get('document_info', {})
        if doc_info:
            doc_fields = {
                'Тип документа': doc_info.get('document_type'),
                'Номер документа': doc_info.get('document_number'),
                'Дата документа': doc_info.get('document_date'),
                'Дата (нормалізована)': doc_info.get('document_date_normalized'),
                'Місце складання': doc_info.get('location'),
                'Валюта': doc_info.get('currency'),
            }
            for field_name, value in doc_fields.items():
                if value is not None:
                    ws1[f'A{row}'] = field_name
                    ws1[f'B{row}'] = str(value)
                    ws1[f'B{row}'].alignment = value_alignment
                    row += 1

        row += 1  # Пустая строка

        # 2. Постачальник
        ws1.merge_cells(f'A{row}:B{row}')
        ws1[f'A{row}'] = "Постачальник"
        ws1[f'A{row}'].fill = section_fill
        ws1[f'A{row}'].font = section_font
        row += 1

        parties = data.get('parties', {})
        supplier = parties.get('supplier', {})
        if supplier:
            supplier_fields = {
                'Назва': supplier.get('name'),
                'Адреса': supplier.get('address'),
                'Телефон': supplier.get('phone'),
                'ЄДРПОУ': supplier.get('tax_id'),
                'ІПН': supplier.get('vat_id'),
                'Номер рахунку': supplier.get('account_number'),
                'Банк': supplier.get('bank'),
            }
            for field_name, value in supplier_fields.items():
                if value is not None:
                    ws1[f'A{row}'] = field_name
                    ws1[f'B{row}'] = str(value)
                    ws1[f'B{row}'].alignment = value_alignment
                    row += 1

        row += 1  # Пустая строка

        # 3. Покупець
        ws1.merge_cells(f'A{row}:B{row}')
        ws1[f'A{row}'] = "Покупець"
        ws1[f'A{row}'].fill = section_fill
        ws1[f'A{row}'].font = section_font
        row += 1

        customer = parties.get('customer', {})
        if customer:
            customer_fields = {
                'Назва': customer.get('name'),
                'Адреса': customer.get('address'),
                'ЄДРПОУ': customer.get('tax_id'),
                'ІПН': customer.get('vat_id'),
            }
            for field_name, value in customer_fields.items():
                if value is not None:
                    ws1[f'A{row}'] = field_name
                    ws1[f'B{row}'] = str(value)
                    ws1[f'B{row}'].alignment = value_alignment
                    row += 1

        row += 1  # Пустая строка

        # 4. Договір та замовлення
        references = data.get('references', {})
        if references:
            ws1.merge_cells(f'A{row}:B{row}')
            ws1[f'A{row}'] = "Договір та замовлення"
            ws1[f'A{row}'].fill = section_fill
            ws1[f'A{row}'].font = section_font
            row += 1

            if 'contract' in references and references['contract'].get('value'):
                ws1[f'A{row}'] = 'Договір'
                ws1[f'B{row}'] = str(references['contract']['value'])
                ws1[f'B{row}'].alignment = value_alignment
                row += 1

            if 'order' in references and references['order'].get('value'):
                ws1[f'A{row}'] = 'Замовлення'
                ws1[f'B{row}'] = str(references['order']['value'])
                ws1[f'B{row}'].alignment = value_alignment
                row += 1

        row += 1  # Пустая строка

        # 5. Підсумки
        totals = data.get('totals', {})
        if totals:
            ws1.merge_cells(f'A{row}:B{row}')
            ws1[f'A{row}'] = "Підсумки"
            ws1[f'A{row}'].fill = section_fill
            ws1[f'A{row}'].font = section_font
            row += 1

            totals_fields = {
                'Сума без ПДВ': totals.get('subtotal'),
                'ПДВ': totals.get('vat'),
                'Всього': totals.get('total'),
            }
            for field_name, value in totals_fields.items():
                if value is not None:
                    ws1[f'A{row}'] = field_name
                    ws1[f'B{row}'] = str(value)
                    row += 1

        # Настройка ширины колонок
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 60

        # Items sheet
        items_sheet_name = self.config.excel_sheet_items_name
        ws2 = wb.create_sheet(items_sheet_name)

        # Заголовки для позиций
        table_data = data.get('table_data', {})
        line_items = table_data.get('line_items', [])
        column_mapping = table_data.get('column_mapping', {})

        if line_items:
            # Определяем колонки на основе column_mapping и данных
            all_keys = set()
            for item in line_items:
                all_keys.update(item.keys())

            # Сортируем ключи: сначала стандартные, потом остальные
            standard_order = ['no', 'line_number', 'name', 'tovar', 'product', 'sku', 'unit',
                            'quantity', 'kilkist', 'price', 'tsina', 'amount', 'suma',
                            'vat_rate', 'vat_amount', 'ukt_zed']
            ordered_keys = []
            for key in standard_order:
                if key in all_keys:
                    ordered_keys.append(key)
            for key in sorted(all_keys):
                if key not in ordered_keys and key not in ['raw', '_meta']:
                    ordered_keys.append(key)

            # Заголовки
            col_idx = 1
            header_row = 1
            for key in ordered_keys:
                # Используем column_mapping для украинских названий
                header_text = column_mapping.get(key, key.replace('_', ' ').title())
                ws2.cell(row=header_row, column=col_idx, value=header_text)
                ws2.cell(row=header_row, column=col_idx).fill = header_fill
                ws2.cell(row=header_row, column=col_idx).font = header_font
                col_idx += 1

            # Данные
            for row_idx, item in enumerate(line_items, start=2):
                col_idx = 1
                for key in ordered_keys:
                    value = item.get(key)
                    if value is not None:
                        # Для числовых значений сохраняем как число
                        if isinstance(value, (int, float)):
                            ws2.cell(row=row_idx, column=col_idx, value=value)
                        else:
                            ws2.cell(row=row_idx, column=col_idx, value=str(value))
                    col_idx += 1

            # Автоподбор ширины колонок
            for col_idx in range(1, len(ordered_keys) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, len(line_items) + 2):
                    cell_value = ws2.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws2.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def _export_simple(self, wb: Workbook, invoice_data: InvoiceData):
        """
        Простой экспорт (fallback)

        Args:
            wb: Workbook
            invoice_data: Данные счета
        """
        # Header sheet
        header_sheet_name = self.config.excel_sheet_header_name
        ws1 = wb.create_sheet(header_sheet_name)
        field_column = self.config.excel_header_field_column
        value_column = self.config.excel_header_value_column
        ws1['A1'], ws1['B1'] = field_column, value_column
        header_dict = invoice_data.header.model_dump()
        row = 2
        for key, value in header_dict.items():
            ws1[f'A{row}'] = key
            ws1[f'B{row}'] = str(value) if value is not None else ""
            row += 1

        # Items sheet
        items_sheet_name = self.config.excel_sheet_items_name
        ws2 = wb.create_sheet(items_sheet_name)
        if invoice_data.items:
            headers = list(invoice_data.items[0].model_dump().keys())
            for col_idx, header in enumerate(headers, start=1):
                ws2.cell(row=1, column=col_idx, value=header)
            for row_idx, item in enumerate(invoice_data.items, start=2):
                item_dict = item.model_dump()
                for col_idx, header in enumerate(headers, start=1):
                    value = item_dict.get(header)
                    ws2.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
